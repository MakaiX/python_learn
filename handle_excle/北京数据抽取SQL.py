# -*- coding: utf-8 -*-

import os
import openpyxl

os.chdir('/Volumes/ShareFiles/3银保监数据标准化/保险业监管数据标准化规范')

wb = openpyxl.load_workbook('附件5：保险业监管数据标准化规范数据结构一览表.xlsx')

# 获取workbook中所有的表格
sheets = wb.sheetnames

# print(sheets)
result_sql = ''
# for i in range(3, len(sheets)):
for i in [3, 10, 11, 12, 13, 14, 45, 48]:
    sheet = wb[sheets[i]]
    table_sql = ''
    table_name = ''
    for r in range(1, sheet.max_row + 1):
        line = ''

        if r == 1:
            strval = str(sheet.cell(row=r, column=1).value).ljust(17)
            import re

            commont = re.findall(r'[^（）]+', strval)[0]
            table_name = re.findall(r'[^（）]+', strval)[1]
            # print(table_name)
        elif r > 2:
            line = str(sheet.cell(row=r, column=2).value).ljust(1) + ' ,'
        table_sql += line  # + '\n'

    table_sql = table_sql.strip(',')
    result_sql += 'SELECT ' + table_sql + 'FROM ' + table_name + ' WHERE SJBSPCH =\'20201231\';' + table_name + ';000166-' + table_name + '-20201231.txt' + '\n'
print(result_sql)
# with open('GeneralSQL.sql', 'a', encoding='utf-8') as f:
#     f.write(result_sql)
