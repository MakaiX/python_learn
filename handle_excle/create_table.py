# -*- coding: utf-8 -*-

import os
import openpyxl

os.chdir('/Volumes/ShareFiles/3银保监数据标准化/保险业监管数据标准化规范')

wb = openpyxl.load_workbook('附件5：保险业监管数据标准化规范数据结构一览表.xlsx')

# 获取workbook中所有的表格
sheets = wb.sheetnames
# print(sheets)
result_sql = ''
for i in range(3, len(sheets)):
    # for i in range(3,4):
    sheet = wb[sheets[i]]
    table_sql = ''
    commont = '--' + sheet.title + '\n'
    table_sql += commont
    pkstr = ''
    table_name = ''
    for r in range(1, sheet.max_row + 1):
        line = ''

        if r == 1:
            # print('\n' + ''.join(
            # [str(sheet.cell(row=r, column=c).value).ljust(17) for c in range(1, sheet.max_column + 1)]))
            # print(str(sheet.cell(row=r, column=1).value).ljust(17))
            strval = str(sheet.cell(row=r, column=1).value).ljust(17)
            import re

            commont = re.findall(r'[^（）]+', strval)[0]
            table_name = re.findall(r'[^（）]+', strval)[1]
            table_sql += 'create table ' + table_name + '('
            print(table_name)
        elif r > 2:
            # LSH	VARCHAR2(28)	not null	,
            # print(str(sheet.cell(row=r, column=5).value).ljust(17))
            if 'N' == str(sheet.cell(row=r, column=5).value).ljust(17).strip():
                line = str(sheet.cell(row=r, column=2).value).ljust(17) + str(sheet.cell(row=r, column=6).value).ljust(
                    17) + ' not null	,'
            else:
                line = str(sheet.cell(row=r, column=2).value).ljust(17) + str(sheet.cell(row=r, column=6).value).ljust(
                    17) + ' ,'

            if 'Y' == str(sheet.cell(row=r, column=4).value).ljust(17).strip():
                pkstr += str(sheet.cell(row=r, column=2).value).ljust(17).strip() + ','

        table_sql += line + '\n'

    pkstr = pkstr.strip(',')
    table_sql += 'SJBSPCH' + '\t'+'Varchar2(10) ,' + '\n'
    table_sql += 'constraint PK_' + table_name.strip() + ' primary key(' + pkstr + ')' + '\n);'
    print(table_sql)
    result_sql += table_sql + '\n'
with open('GeneralSQL.sql', 'a', encoding='utf-8') as f:
    f.write(result_sql)
